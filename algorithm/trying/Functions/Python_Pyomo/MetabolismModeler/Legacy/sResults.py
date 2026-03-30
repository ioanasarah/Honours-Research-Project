import os
import pandas as pd
import numpy as np
import pyomo.environ as pe
import cobra
import openpyxl
from openpyxl.utils import get_column_letter

__all__ = [
    "save_solver_results",
    "get_active_reactions_from_solved_model",
    "save_excel_or_CSV_if_file_already_exist",
    "find_or_create_user_based_directory",
]


def find_or_create_user_based_directory(
    output_folder: str, user_name: str, specific_folder=None, base_files: bool = False
) -> (str, str):
    # assume program is ran from a folder SRC inside a repository.
    # the SRC is inside Git\\Metabolic_Task_Score\\Data\\pyomo_python_files
    # the main_data_folder is then: Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\INIT_runs_for_testing
    # user folder for output is then: Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\user_name\\output_folder
    temporary_path = os.getcwd()
    try:
        main_data_folder = os.path.abspath(
            os.path.join(os.getcwd(), "..", "..", "..", "Data", "pyomo_python_files")
        )
        user_folder = os.path.join(main_data_folder, user_name)
        os.makedirs(user_folder, exist_ok=True)
        output_folder = os.path.join(user_folder, output_folder)
        os.makedirs(output_folder, exist_ok=True)
        if base_files:
            main_data_folder = os.path.abspath(
                os.path.join(
                    os.getcwd(),
                    "..",
                    "..",
                    "..",
                    "Data",
                    "pyomo_python_files",
                    "Base_files",
                )
            )
        elif specific_folder is not None:
            main_data_folder = os.path.join(main_data_folder, specific_folder)
    except Exception as e:
        print(e)
        main_data_folder = temporary_path
        output_folder = temporary_path
    return main_data_folder, output_folder


def save_excel_or_CSV_if_file_already_exist(df, filename_location):
    counter = 1
    filename_without_extension = os.path.splitext(filename_location)[0]
    saved_file = False
    try:
        filename = f"{filename_without_extension}.xlsx"
        if not os.path.exists(filename):
            df.to_excel(filename, index=False, header=False)
            wb = openpyxl.load_workbook(filename)
            wb.calculation.calculation_on_save = False
            # set width to be 10 or length of longest string in column at max 30
            for idx, column in enumerate(df.columns, start=1):
                max_length = max(
                    df[column]
                    .map(
                        lambda x: (
                            len(str(round(x, 2)))
                            if isinstance(x, float)
                            else len(str(x))
                        )
                    )
                    .max(),
                    0,
                )
                column_letter = get_column_letter(
                    idx
                )  # Convert column index to Excel column letter
                # max_length = max(df[column].map(str).map(len).max(), 0)
                # column_letter = get_column_letter(idx)  # Convert column index to Excel column letter
                wb.active.column_dimensions[column_letter].width = min(
                    max_length + 3, 50
                )

            wb.save(filename)
            wb.close()
            saved_file = True
    except Exception as e:
        print(e)
        filename = f"{filename_without_extension}.csv"
        if not os.path.exists(filename):
            df.to_csv(filename, index=False, header=False)
            saved_file = True

    while not saved_file:
        try:
            filename_with_counter = f"{filename_without_extension}_Try{counter}.xlsx"
            if not os.path.exists(filename_with_counter):
                df.to_excel(filename_with_counter, index=False, header=False)
                wb = openpyxl.load_workbook(filename_with_counter)
                # set width to be 10 or length of longest string in column at max 30
                for idx, column in enumerate(df.columns, start=1):
                    max_length = max(
                        df[column]
                        .map(
                            lambda x: (
                                len(str(round(x, 2)))
                                if isinstance(x, float)
                                else len(str(x))
                            )
                        )
                        .max(),
                        0,
                    )
                    column_letter = get_column_letter(
                        idx
                    )  # Convert column index to Excel column letter
                    # max_length = max(df[column].map(str).map(len).max(), 0)
                    # column_letter = get_column_letter(idx)  # Convert column index to Excel column letter
                    wb.active.column_dimensions[column_letter].width = min(
                        max_length + 3, 50
                    )

                wb.save(filename_with_counter)
                wb.close()

                break
            else:
                counter += 1
        except Exception as e:
            print(e)
            filename_with_counter = f"{filename_without_extension}_Try{counter}.csv"
            if not os.path.exists(filename_with_counter):
                df.to_csv(filename_with_counter, index=False, header=False)
                break
            else:
                counter += 1


def save_solver_results(
    instance: pe.ConcreteModel,
    result,
    cobra_model: cobra.Model,
    filename: str,
    main_folder: str,
    expression_array: np.ndarray = None,
    old_expression_array: np.ndarray = None,
    epsilon: float = 1e-3,
) -> None:
    var_values = [
        (v.name, v.value) for v in instance.component_data_objects(pe.Var, active=True)
    ]
    var_values = [
        (name.split("[")[0], int(name.split("[")[1].split("]")[0]), value)
        for name, value in var_values
    ]
    df_vars = pd.DataFrame(var_values, columns=["Variable", "Index", "Value"])

    min_indices = df_vars.groupby("Variable")["Index"].min().reset_index()
    min_indices.columns = ["Variable", "MinIndex"]
    df_vars = df_vars.merge(min_indices, on="Variable")
    df_vars["AdjustedIndex"] = df_vars.apply(
        lambda row: row["Index"] + 1 if row["MinIndex"] == 0 else row["Index"], axis=1
    )
    df_vars = df_vars.pivot(
        index="AdjustedIndex", columns="Variable", values="Value"
    ).reset_index()
    # df_vars = df_vars.pivot(index='Index', columns='Variable', values='Value').reset_index()

    if cobra_model is not None:
        df_vars["Reaction Name"] = [reaction.name for reaction in cobra_model.reactions]
        df_vars["Reaction ID"] = [reaction.id for reaction in cobra_model.reactions]
        df_vars["Reaction Formula"] = [
            str(reaction.build_reaction_string(use_metabolite_names=True))
            for reaction in cobra_model.reactions
        ]
        df_vars["Reaction Formula Using IDs"] = [
            str(reaction.build_reaction_string(use_metabolite_names=False))
            for reaction in cobra_model.reactions
        ]
    if expression_array is not None:
        df_vars["Expression"] = old_expression_array
        df_vars["Modified_Expression"] = expression_array
    df_vars.drop(columns=["AdjustedIndex"], inplace=True)
    df_vars.insert(0, "Zeros", 0)

    columns_order = [
        "Zeros",
        "V_flux",
        "Expression",
        "Reaction ID",
        "Reaction Name",
        "Reaction Formula",
        "Reaction Formula Using IDs",
    ]
    other_columns = [col for col in df_vars.columns if col not in columns_order]
    df_vars = df_vars[columns_order + other_columns]

    result_dict = {
        "Termination condition": result.Solver.Termination_condition,
        "Time": result.Solver.Time,
        "Objective": instance.objective(),
    }
    # df_result = pd.DataFrame.from_records([result_dict]).transpose()
    df_result = pd.DataFrame(list(result_dict.items()), columns=["Settings", "Values"])
    # df_settings = pd.DataFrame(df_result[0].values, columns=['Settings'])
    # df_settings = df_settings.reset_index(drop =True)
    # df_result.reset_index(drop = True)
    df_settings = df_result

    df_final = pd.concat([df_vars, df_settings], axis=1)

    location_to_save = os.path.join(main_folder, filename)
    save_excel_or_CSV_if_file_already_exist(df_final, location_to_save)

    columns_to_check = ["V", "Y", "V_flux", "Y_rxn_active"]

    for column in columns_to_check:
        if column in df_vars.columns:
            df_filtered = df_vars[abs(df_vars[column]) >= epsilon]
            df_filtered = df_filtered.reset_index(drop=True)
            df_final_filtered = pd.concat([df_filtered, df_settings], axis=1)
            location_to_save_filtered = os.path.join(
                main_folder, f"filtered_{filename}"
            )
            save_excel_or_CSV_if_file_already_exist(
                df_final_filtered, location_to_save_filtered
            )
            break


if __name__ == "__main__":
    placeholder = 5
    # main_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\INIT_runs_for_testing"
    # save_solver_results(instance, result, 'solver_results.csv', main_folder)


def get_active_reactions_from_solved_model(
    instance,
    expression_old: [],
    expression: [],
    cobra_model: cobra.Model,
    **settings_kwargs_dict: dict,
) -> (dict, tuple):
    reactions_included_this_task = (
        {}
    )  # task_number: {reaction_id: (reaction_name, flux, reaction_formula)}
    if "epsilon" in settings_kwargs_dict:
        epsilon = settings_kwargs_dict["epsilon"]
    else:
        epsilon = 1e-3
    var_values = [
        (v.name, v.value) for v in instance.component_data_objects(pe.Var, active=True)
    ]
    var_values = [
        (name.split("[")[0], int(name.split("[")[1].split("]")[0]), value)
        for name, value in var_values
    ]
    df_vars = pd.DataFrame(var_values, columns=["Variable", "Index", "Value"])
    df_vars = df_vars.pivot(
        index="Index", columns="Variable", values="Value"
    ).reset_index()
    for rxn in range(1, len(cobra_model.reactions) + 1):
        if abs(df_vars["V_flux"][rxn - 1]) >= epsilon:
            name = cobra_model.reactions[rxn - 1].name
            id = cobra_model.reactions[rxn - 1].id
            formula = cobra_model.reactions[rxn - 1].build_reaction_string(
                use_metabolite_names=True
            )
            flux = df_vars["V_flux"][rxn - 1]
            reactions_included_this_task[id] = (
                0,
                flux,
                expression_old[rxn - 1],
                id,
                name,
                formula,
                expression[rxn - 1],
            )
    amount_of_reactions = len(reactions_included_this_task.keys())
    amount_of_expressionless_reactions = len(
        [value for key, value in reactions_included_this_task.items() if value[2] == -1]
    )
    amount_of_reactions_with_expression = (
        amount_of_reactions - amount_of_expressionless_reactions
    )
    print(
        f"\nAmount of reactions with expression: {amount_of_reactions_with_expression}"
    )
    print(
        f"Amount of reactions without expression: {amount_of_expressionless_reactions}"
    )
    print(f"Total amount of reactions: {amount_of_reactions}")
    return reactions_included_this_task, (
        amount_of_reactions,
        amount_of_expressionless_reactions,
        amount_of_reactions_with_expression,
    )
