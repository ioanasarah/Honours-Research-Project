import json
import os
import re
import glob
import shutil
import cobra.io
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from json import load
from warnings import catch_warnings, filterwarnings
from logging import getLogger, ERROR
from scipy.io import loadmat
from cobra.io import load_matlab_model, save_matlab_model
from cobra import Model, Reaction, Metabolite
from hdf5storage import loadmat

__all__ = [
    "find_all_folders_with_excel_file",
    "import_necessary_dot_Mat_files_from_task_directories",
    "extract_info_from_log_files",
    "create_pd_dataframe_from_dict",
    "rename_tasks_to_n_digits",
    "save_excel_or_CSV_if_file_already_exist",
]


def load_json_data(main_folder: str,
                   expression_type: str = "local_threshold",
                   ) -> (list, list, list):
    if expression_type == "local_threshold":
        try:
            expression = load(open(main_folder + "\\" + "expressionRxns.json", "r"))
        except:
            expression = load(open(main_folder + "\\" + "expression_reversible.json", "r"))

    elif expression_type == "global_no_threshold":
        try:
            expression = load(open(main_folder + "\\" + "expressionRxns_no_threshold.json", "r"))
        except:
            expression = load(open(main_folder + "\\" + "expression_reversible_no_threshold.json", "r"))

    try:
        significance = load(open(main_folder + "\\" + "significance.json", "r"))
    except:
        significance = load(
            open(main_folder + "\\" + "significance_reversible.json", "r")
        )
    try:
        task_structure = load(open(main_folder + "\\" + "taskStructure.json", "r"))
    except:
        task_structure = load(open(main_folder + "\\" + "task_structure.json", "r"))

    return expression, significance, task_structure


def find_or_create_user_based_directory(
    output_folder: str, user_name: str, specific_folder=None, base_files: bool = False
) -> (str, str):
    # assume program is ran from a folder SRC inside a repository.
    # the SRC is inside Git\\Metabolic_Task_Score\\Data\\pyomo_python_files
    # the main_data_folder is then: Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\INIT_runs_for_testing
    # user folder for output is then: Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\user_name\\output_folder
    temporary_path = os.getcwd()
    try:
        main_data_folder = os.path.abspath(
            os.path.join(os.getcwd(), "..", "..", "..", "Data", "pyomo_python_files")
        )
        user_folder = os.path.join(main_data_folder, user_name)
        os.makedirs(user_folder, exist_ok=True)
        output_folder = os.path.join(user_folder, output_folder)
        os.makedirs(output_folder, exist_ok=True)
        if base_files:
            main_data_folder = os.path.abspath(
                os.path.join(
                    os.getcwd(),
                    "..",
                    "..",
                    "..",
                    "Data",
                    "pyomo_python_files",
                    "Base_files",
                )
            )
        elif specific_folder is not None:
            main_data_folder = os.path.join(main_data_folder, specific_folder)
    except Exception as e:
        print(e)
        main_data_folder = temporary_path
        output_folder = temporary_path
    return main_data_folder, output_folder


def load_base_files_in_base_directory(path):
    files_to_load = [
        "expressionRxnsCellfie.mat"
    ]  # , "model.mat", "significance.mat", "taskData.mat"]
    model = None
    expression = None
    significance = None
    task_data = None
    try:
        main_data_folder = os.path.abspath(
            os.path.join(os.getcwd(), "..", "..", "..", "Data", "pyomo_python_files")
        )
        base_files_folder = os.path.join(main_data_folder, "Base_files")
        os.makedirs(base_files_folder, exist_ok=True)
    except Exception as e:
        print(e)
        print("Could not locate base files folder")
        return
    for file_to_get in files_to_load:
        if path is not None:
            file = os.path.join(path, file_to_get)
        else:
            file = os.path.join(base_files_folder, file_to_get)
        print(f"Loading {file_to_get}")
        if file_to_get == "model.mat":
            with catch_warnings():
                original_logging_level = getLogger().getEffectiveLevel()
                getLogger().setLevel(ERROR)
                filterwarnings("ignore", "No defined compartments in model")
                model = load_matlab_model(file)
                getLogger().setLevel(original_logging_level)
        elif file_to_get == "expressionRxns.mat":
            expression = loadmat(file)
            # expression_cells = scipy.io.loadmat(file)
            # expression = [value[0] for value in expression_cells['expressionRxns']]
        elif file_to_get == "significance.mat":
            signficance_cels = loadmat(file)
            # expression = [value[0] for value in expression_cells['significance']]
        elif file_to_get == "taskData.mat":
            task_data = loadmat(file)
            # expression = [value[0] for value in expression_cells['taskData']]
        print(f"{file_to_get} loaded successfully")

    return model, expression, significance, task_data

def prepare_task_structure_to_correct_format(task_structure, output_folder):
    pass
    ### TODO this code is not yet implemented, is present in sResults load_json
    ### the task structure had a weird structure and I manually adapted it once,
    # however it should be created for full integration of the code
    # json.dump(task_structure, open(output_folder + "\\" + "taskStructure.json", "w"))


def fix_and_save_loaded_mat_files(expression, significance, task_structure, model, output_folder):
    expression = [list(row) for row in expression["expressionRxns"]]
    significance = [list(row) for row in significance["significance"]]
    prepare_task_structure_to_correct_format(task_structure, output_folder)

    json.dump(expression, open(output_folder + "\\" + "expressionRxns.json", "w"))
    json.dump(significance, open(output_folder + "\\" + "significance.json", "w"))
    cobra.io.save_matlab_model(model, output_folder + "\\" + "modelCellfie.mat")


def load_and_adapt_mat_files_for_python(
    main_folder_with_mat_files: str, user_name: str, specific_folder=None
):
    main_data_folder, output_folder = find_or_create_user_based_directory(
        "test_runs", user_name, specific_folder, base_files=True
    )
    model, expression, significance, task_structure = load_base_files_in_base_directory(
        main_folder_with_mat_files
    )
    fix_and_save_loaded_mat_files(expression, significance, task_structure, model, output_folder)

if __name__ == "__main__":
    placeholder = 5
    # main_folder_1 = "E:\\Git\\Metabolic_Task_Score\\Data\\Base MAT Files\\Consensus Model 1.17\\No Thresh\\"
    # model, expression, significance, task_structure = load_base_files_in_base_directory(main_folder_1)

    # # print(model)
    # print(expression["expressionRxns"])
    # os.getcwd()
    # import json
    # expression = [list(row) for row in expression["expressionRxns"]]
    # json.dump(expression, open("expressionRxns_no_threshold.json", "w"))
    # print(significance)
    # print(task_data)


def obtain_cobra_model_from_matlab_workspace(
    matfile: str, model_name: str, output_folder: str
):
    print("This function can take considerable amount of time (15 min)")
    matfile = loadmat(matfile, squeeze_me=True)
    model_struct = matfile[model_name]

    # Extract fields
    S = model_struct["S"]
    lb = model_struct["lb"]
    ub = model_struct["ub"]
    c = model_struct["c"]
    b = model_struct["b"]
    genes = model_struct["genes"]
    rxns = model_struct["rxns"]
    mets = model_struct["mets"]
    rules = model_struct["rules"]
    compNames = model_struct["compNames"]
    comps = model_struct["comps"]
    metCharges = model_struct["metCharges"]
    metFormulas = model_struct["metFormulas"]
    metNames = model_struct["metNames"]
    grRules = model_struct["grRules"]
    rxnGeneMat = model_struct["rxnGeneMat"]
    rxnNames = model_struct["rxnNames"]
    rxnNotes = model_struct["rxnNotes"]
    rxnReferences = model_struct["rxnReferences"]
    subSystems = model_struct["subSystems"]
    csense = model_struct["csense"]
    osenseStr = model_struct["osenseStr"]
    metInChIString = model_struct["metInChIString"]
    rxnECNumbers = model_struct["rxnECNumbers"]
    rxnConfidenceScores = model_struct["rxnConfidenceScores"]
    description = model_struct["description"]
    modelID = model_struct["modelID"]
    version = model_struct["version"]
    # Convert to more accessible format if necessary
    S = S.item() if hasattr(S, "item") else S
    lb = lb.item() if hasattr(lb, "item") else lb
    ub = ub.item() if hasattr(ub, "item") else ub
    c = c.item() if hasattr(c, "item") else c
    genes = genes.item() if hasattr(genes, "item") else genes
    rxns = rxns.item() if hasattr(rxns, "item") else rxns
    mets = mets.item() if hasattr(mets, "item") else mets
    rules = rules.item() if hasattr(rules, "item") else rules
    compNames = compNames.item() if hasattr(compNames, "item") else compNames
    comps = comps.item() if hasattr(comps, "item") else comps
    metCharges = metCharges.item() if hasattr(metCharges, "item") else metCharges
    metFormulas = metFormulas.item() if hasattr(metFormulas, "item") else metFormulas
    metNames = metNames.item() if hasattr(metNames, "item") else metNames
    grRules = grRules.tolist()
    rxnGeneMat = rxnGeneMat.item() if hasattr(rxnGeneMat, "item") else rxnGeneMat
    rxnNames = rxnNames.item() if hasattr(rxnNames, "item") else rxnNames
    rxnNotes = rxnNotes.item() if hasattr(rxnNotes, "item") else rxnNotes
    rxnReferences = (
        rxnReferences.item() if hasattr(rxnReferences, "item") else rxnReferences
    )
    subSystems = subSystems.item() if hasattr(subSystems, "item") else subSystems
    csense = csense.item() if hasattr(csense, "item") else csense
    osenseStr = osenseStr.item() if hasattr(osenseStr, "item") else osenseStr
    metInChIString = (
        metInChIString.item() if hasattr(metInChIString, "item") else metInChIString
    )
    rxnECNumbers = (
        rxnECNumbers.item() if hasattr(rxnECNumbers, "item") else rxnECNumbers
    )
    rxnConfidenceScores = (
        rxnConfidenceScores.item()
        if hasattr(rxnConfidenceScores, "item")
        else rxnConfidenceScores
    )
    description = description.item() if hasattr(description, "item") else description
    modelID = modelID.item() if hasattr(modelID, "item") else modelID
    version = version.item() if hasattr(version, "item") else version

    cobra_model = Model("Base_model")
    # Add metabolites
    metabolites = []
    for met in mets:
        metabolite = Metabolite(met)
        metabolites.append(metabolite)
    cobra_model.add_metabolites(metabolites)
    for i, metabolite in enumerate(cobra_model.metabolites):
        metabolite.charge = metCharges[i] if metCharges[i] is not None else 0
        metabolite.formula = metFormulas[i] if len(metFormulas[i]) > 0 else ""
        metabolite.name = metNames[i] if len(metNames[i]) > 0 else ""
        # metabolite.inchi = metInChIString[i] if len(metInChIString[i]) > 0 else ''

    # Add reactions
    reactions = [0] * len(rxns)
    for i, rxn_id in enumerate(rxns):
        reaction = Reaction(rxn_id)
        print(f"Adding reaction {rxn_id}")
        # if i == 20:
        #     break
        # cobra_model.add_reactions([reaction])  # Add reaction to the model first
        reaction.lower_bound = lb[i]
        reaction.upper_bound = ub[i]
        for j, met in enumerate(mets):
            coeff = S[j, i]
            if coeff != 0:
                reaction.add_metabolites(
                    {cobra_model.metabolites.get_by_id(met): coeff}
                )
        reactions[i] = reaction

    cobra_model.add_reactions(reactions)
    for i, reaction in enumerate(cobra_model.reactions):
        # Set gene-reaction rules (grRules)
        reaction.gene_reaction_rule = grRules[i] if len(grRules[i]) > 0 else ""

        reaction.objective_coefficient = c[i]  # Now set the objective coefficient
        # Set reaction references (rxnReferences)
        reaction.references = rxnReferences[i] if len(rxnReferences[i]) > 0 else ""
        # reaction.notes = rxnNotes[i] if len(rxnNotes[i]) > 0 else ''
        # reaction.subsystem = subSystems[i] if len(subSystems[i]) > 0 else ''
        reaction.name = rxnNames[i] if len(rxnNames[i]) > 0 else ""

        # Set subsystems
        if type(subSystems[i]) == str:
            reaction.subsystem = subSystems[i] if len(subSystems[i]) > 0 else ""
        else:
            reaction.subsystem = (
                str(subSystems[i].tolist()) if len(subSystems[i]) > 0 else ""
            )

    output_file = os.path.join(output_folder, "model_loaded.mat")
    save_matlab_model(cobra_model, output_file)


if __name__ == "__main__":
    placeholder = 5
    # output_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\Base_files"
    # model_name = "model"
    # matfile = "E:\\Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\Base_files\\matlab_workspace.mat"
    # obtain_cobra_model_from_matlab_workspace(matfile, model_name, output_folder)


def find_all_folders_with_excel_file(
    main_folder: str, total_number_of_tasks: int
) -> (list, list):
    def check_directories_for_excel(main_folder, max_task_number):
        excel_extensions = ["*.xlsx", "*.xls"]

        task_set = list(f"Task{num}" for num in range(1, max_task_number + 1))
        existing_tasks_with_excel = []

        for root, dirs, files in os.walk(main_folder):
            for idx, task in enumerate(task_set):
                pattern = rf"Task{idx + 1}S"
                for dir_ in dirs:
                    if re.search(pattern, dir_):
                        for extension in excel_extensions:
                            search_pattern = os.path.join(root, "**", extension)
                            if glob.glob(search_pattern, recursive=True):
                                existing_tasks_with_excel.append(idx)
                                break
                        break
            break
        tasks_without_excel = [
            number
            for number in range(1, max_task_number)
            if number not in existing_tasks_with_excel
        ]

        return existing_tasks_with_excel, tasks_without_excel

    # Example usage:
    # main_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\INIT_Like_RUNS"
    # total_number_of_tasks = 325
    folders_with_excel, tasks_without_excel = check_directories_for_excel(
        main_folder, total_number_of_tasks
    )
    print("Tasks done:")
    print(folders_with_excel)
    print("\nTasks not yet done or did not converge before abandoning:")
    print(tasks_without_excel)
    return (folders_with_excel, tasks_without_excel)


###
def import_necessary_dot_Mat_files_from_task_directories(
    task_list: list, main_folder: str, output_folder: str
) -> None:
    filenames_to_get = [
        "expValueMeanAdjustedTaskIrrevKapprox.mat",
        "expValueMeanAdjustedTaskReversibleKApproximation.mat",
        "newIrrevModelKapprox.mat",
        "newReversibleModelForKApproxmiation.mat",
        "rev2irrevIrrevKapprox.mat",
        "logFile.txt",
        "ActiveReactions",
    ]

    all_dirs = [dirpath for dirpath, dirnames, files in os.walk(main_folder)]
    for task_num in task_list:
        pattern = rf"Task{task_num}S"
        for dir_ in all_dirs:
            if re.search(pattern, dir_):
                # create new folder in output folder of same name as current dir_
                # create new folder in output folder of same name as current dir_
                task_output_folder = os.path.join(output_folder, os.path.basename(dir_))
                new_name = re.sub(
                    r"Task(\d+)S",
                    lambda m: f"Task{int(m.group(1)):0{3}}S",
                    task_output_folder,
                )
                os.makedirs(new_name, exist_ok=True)
                task_output_folder = new_name
                for file_to_get in filenames_to_get:
                    if file_to_get == "ActiveReactions":
                        search_pattern = os.path.join(
                            dir_, "**", "*ActiveReactions*.xlsx"
                        )
                        files = glob.glob(search_pattern, recursive=True)
                        if files:
                            for file in files:
                                destination_file = os.path.join(
                                    task_output_folder, os.path.basename(file)
                                )
                                if not os.path.isfile(destination_file):
                                    os.system(f"copy {file} {destination_file}")

                    search_pattern = os.path.join(dir_, "**", file_to_get)
                    files = glob.glob(search_pattern, recursive=True)
                    if files:
                        for file in files:
                            destination_file = os.path.join(
                                task_output_folder, os.path.basename(file)
                            )
                            if not os.path.isfile(destination_file):
                                # print(f"Copying {file} to {task_output_folder}")
                                os.system(f"copy {file} {destination_file}")
                break


if __name__ == "__main__":
    placeholder = 5
    # task_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
    # task_list = [11, 6, 15, 24, 25, 88, 93, 85, 1, 96, 86, 2, 3, 4, 5, 84, 10,251, 250]
    # main_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\INIT_Like_RUNS\\"
    # output_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\Pyomo_python_files\\INIT_runs_for_testing\\"
    # import_necessary_dot_Mat_files_from_task_directories(task_list, main_folder, output_folder)


####


def extract_info_from_log_files(main_folder: str) -> dict:
    result_dict = {}
    task_set = list(f"Task{num}" for num in range(1, 500 + 1))
    all_dirs = [dirpath for dirpath, dirnames, files in os.walk(main_folder)]

    for idx, task in enumerate(task_set):
        pattern = rf"Task{idx + 1}S"
        for dir_ in all_dirs:
            if re.search(pattern, dir_):
                for root, dirs, files in os.walk(dir_):
                    if "logFile.txt" in files:
                        with open(os.path.join(root, "logFile.txt"), "r") as log_file:
                            content = log_file.read()
                            time_to_solve = re.search(r"in (\d+.\d+) seconds", content)
                            found_solution = re.search(
                                r"(Optimal|Interrupted)", content
                            )
                            if time_to_solve and found_solution:
                                result_dict[idx + 1] = (
                                    found_solution.group(),
                                    float(time_to_solve.group(1)),
                                )
                break
    return result_dict


def create_pd_dataframe_from_dict(result_dict: dict) -> pd.DataFrame:
    df = pd.DataFrame(result_dict.items(), columns=["Task", "Solution"])
    df["Found Solution"] = df["Solution"].apply(lambda x: x[0])
    df["Time"] = df["Solution"].apply(lambda x: x[1])
    df.drop(columns=["Solution"], inplace=True)
    df.sort_values(by="Time", inplace=True)
    return df


if __name__ == "__main__":
    placeholder = 5
    # main_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\INIT_runs_for_testing"
    # main_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\INIT_Like_RUNS\\"
    # result_dict = extract_info_from_log_files(main_folder)
    # df = create_pd_dataframe_from_dict(result_dict)
    # print(df.to_string(index=False))
    # df.to_csv("E:\\Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\INIT_runs_for_testing\\solution_info.csv", index=False)
    # [11, 6, 15, 24, 25, 88, 93, 85, 1, 96, 86, 2, 3, 4, 5, 84, 10] # test set in ascending order of solving time


def rename_tasks_to_n_digits(main_folder: str, amount_of_digits: int = 3) -> None:
    for root, dirs, files in os.walk(main_folder):
        for name in dirs + files:
            match = re.search(r"Task(\d+)S", name)
            if match:
                new_name = re.sub(
                    r"task(\d+)s",
                    lambda m: f"task{int(m.group(1)):0{amount_of_digits}}s",
                    name,
                )
                os.rename(os.path.join(root, name), os.path.join(root, new_name))


if __name__ == "__main__":
    placeholder = 5
    # main_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\pyomo_python_files\\INIT_runs_for_testing"
    # rename_tasks_to_n_digits(main_folder)


def save_excel_or_CSV_if_file_already_exist(df, filename_location, with_header=False, make_CSV=False):
    counter = 1
    filename_without_extension = os.path.splitext(filename_location)[0]
    saved_file = False
    try:
        if make_CSV:
            filename = f"{filename_without_extension}.csv"
        else:
            filename = f"{filename_without_extension}.xlsx"
        if not os.path.exists(filename):
            if make_CSV:
                df.to_csv(filename, index=False, header=with_header)
            else:
                df.to_excel(filename, index=False, header=with_header)
                wb = openpyxl.load_workbook(filename)
                wb.calculation.calculation_on_save = False
                # set width to be 10 or length of longest string in column at max 30
                for idx, column in enumerate(df.columns, start=1):
                    max_length = max(
                        df[column]
                        .map(
                            lambda x: (
                                len(str(round(x, 2)))
                                if isinstance(x, float)
                                else len(str(x))
                            )
                        )
                        .max(),
                        0,
                    )
                    column_letter = get_column_letter(
                        idx
                    )  # Convert column index to Excel column letter
                    # max_length = max(df[column].map(str).map(len).max(), 0)
                    # column_letter = get_column_letter(idx)  # Convert column index to Excel column letter
                    wb.active.column_dimensions[column_letter].width = min(
                        max_length + 3, 50
                    )

                wb.save(filename)
                wb.close()
            saved_file = True
    except Exception as e:
        print(e)
        filename = f"{filename_without_extension}.csv"
        if not os.path.exists(filename):
            df.to_csv(filename, index=False, header=with_header)
            saved_file = True

    while not saved_file:
        try:
            if make_CSV:
                filename_with_counter = f"{filename_without_extension}_Try{counter}.csv"
            else:
                filename_with_counter = f"{filename_without_extension}_Try{counter}.xlsx"

            if not os.path.exists(filename_with_counter):
                if make_CSV:
                    df.to_csv(filename_with_counter, index=False, header=with_header)
                else:
                    df.to_excel(filename_with_counter, index=False, header=with_header)
                    wb = openpyxl.load_workbook(filename_with_counter)
                    # set width to be 10 or length of longest string in column at max 30
                    for idx, column in enumerate(df.columns, start=1):
                        max_length = max(
                            df[column]
                            .map(
                                lambda x: (
                                    len(str(round(x, 2)))
                                    if isinstance(x, float)
                                    else len(str(x))
                                )
                            )
                            .max(),
                            0,
                        )
                        column_letter = get_column_letter(
                            idx
                        )  # Convert column index to Excel column letter
                        # max_length = max(df[column].map(str).map(len).max(), 0)
                        # column_letter = get_column_letter(idx)  # Convert column index to Excel column letter
                        wb.active.column_dimensions[column_letter].width = min(
                            max_length + 3, 50
                        )

                    wb.save(filename_with_counter)
                    wb.close()

                break
            else:
                counter += 1
        except Exception as e:
            print(e)
            filename_with_counter = f"{filename_without_extension}_Try{counter}.csv"
            if not os.path.exists(filename_with_counter):
                df.to_csv(filename_with_counter, index=False, header=with_header)
                break
            else:
                counter += 1

def rename_accidentally_misnamed_log_files(directory_with_files: str) -> None:
    # "INIT_irrev_Task_136_Sample001_local_threshold_gurobi_seed4.log_Try1_Try2_Try3_Try4_Try5_Try6_Try7_Try8_Try9_Try10_Try11_Try12_Try13_Try14_Try15_Try16_Try17_Try18_Try19_Try20_Try21_Try22_Try23_Try24_Try25_Try26_Try27_Try28_Try29_Try30"
    for file in os.listdir(directory_with_files):
        if not file.endswith(".log") and ".log" in file:
            split_file = file.split(".log")
            last_part = split_file[1]
            split_by_try = last_part.split("_Try")
            last_number = split_by_try[-1]
            if last_number.isdigit():
                new_name = split_file[0] + f"_Try{last_number}" + ".log"
                os.rename(os.path.join(directory_with_files, file), os.path.join(directory_with_files, new_name))




if __name__ == "__main__":
    # E:\Git\Metabolic_Task_Score\Data\Pyomo_python_files\Jelle\test_runs\fractional_fastest_solving_Time0_T136_S001_ELT
    # folder = "E:\\Git\\Metabolic_Task_Score\\Data\\Pyomo_python_files\\Jelle\\test_runs\\fractional_fastest_solving_Time0_T136_S001_ELT"
    # rename_accidentally_misnamed_log_files(folder)
    pass


# Function to organize files
def organize_files(selected_dir):
    folders = [
        'Excel_for_esher',
        'Filtered_excel_files',
        'Gurobi_log_files',
        'Json_models_for_esher',
        'Task_Jsons',
        'Unfiltered_excel_files'
    ]
    folder_map = {
        'esher_': 'Excel_for_esher',
        'filtered_': 'Filtered_excel_files',
        '_gurobi.log': 'Gurobi_log_files',
        'model_json_': 'Json_models_for_esher'
    }
    for folder in folders:
        folder_path = os.path.join(selected_dir, folder)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
    for filename in os.listdir(selected_dir):
        file_path = os.path.join(selected_dir, filename)
        if os.path.isfile(file_path):
            moved = False
            for key, folder in folder_map.items():
                if filename.startswith(key) or filename.endswith(key):
                    shutil.move(file_path, os.path.join(selected_dir, folder, filename))
                    moved = True
                    break
            if not moved and filename.endswith('.json'):
                shutil.move(file_path, os.path.join(selected_dir, 'Task_Jsons', filename))
                moved = True
            if not moved and (filename.endswith('.xlsx') or filename.endswith('.xls')):
                shutil.move(file_path, os.path.join(selected_dir, 'Unfiltered_excel_files', filename))
if __name__ == "__main__":
    # main_folder = "E:\\Git\\Metabolic_Task_Score\\Data\\consensus_no_threshold_fractional_task1_all_samples"
    # organize_files(main_folder)
    pass
